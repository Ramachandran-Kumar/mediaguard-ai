"""
╔══════════════════════════════════════════════════════════════════╗
║  MEDIAGUARD AI — CMS Reference Database Setup                    ║
║  Builds SQLite DB from all CMS reference files — run once        ║
╚══════════════════════════════════════════════════════════════════╝

PURPOSE:
  Reads all five CMS reference files once and loads them into a
  local SQLite database. After this runs, the pipeline queries the
  DB instead of reading flat files every run — milliseconds vs seconds.

RUN ONCE:
  python cms_db_setup.py

OUTPUT:
  data/mediaguard_reference.db  ← SQLite database (~30-50 MB)

TABLES CREATED:
  ncci_bundles    — 131,611 active hard NCCI edit pairs
  mue_limits      — 12,453 CPT MUE limits
  icd10_codes     — 70,000+ ICD-10-CM codes + descriptions
  cpt_rates       — CPT national payment rates
  icd9_crosswalk  — 14,145 ICD-9 → ICD-10 mappings
  icd_cpt_rules   — Domain-curated ICD-10 → valid CPT mappings (FWA validation)

WHY icd_cpt_rules EXISTS AS A SEPARATE TABLE:
  CMS does not publish an ICD-to-CPT mapping file — this is clinical domain
  knowledge curated from coding guidelines and medical necessity criteria.
  Stored in DB (not hardcoded) so it can be queried, extended, and audited.
  Future: mine SynPUF co-occurrence patterns to expand coverage.

THEN USE IN PIPELINE:
  from cms_db_loader import load_all_from_db
  NCCI_BUNDLES, MUE_LIMITS, ICD_REFERENCE, CPT_REFERENCE, ICD9_MAP, ICD_VALID_CPTS = load_all_from_db()

FILES REQUIRED (in data/ folder):
  ccipra-v321r0-f1.xlsx                              ← NCCI PTP edits
  MCR_MUE_PractitionerServices_Eff_04-01-2026.xlsx   ← MUE limits
  Code Descriptions/icd10cm_codes_2026               ← ICD-10-CM codes
  PFREV26B.txt                                       ← Physician Fee Schedule
  2018_I9gem                                         ← ICD-9→ICD-10 crosswalk
"""

import os
import sqlite3
import time
import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────────────
_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
DB_PATH   = os.path.join(_DATA_DIR, "mediaguard_reference.db")

NCCI_FILE    = os.path.join(_DATA_DIR, "ccipra-v321r0-f1.xlsx")
MUE_FILE     = os.path.join(_DATA_DIR, "MCR_MUE_PractitionerServices_Eff_04-01-2026.xlsx")
ICD10_FILE   = os.path.join(_DATA_DIR, "Code Descriptions", "icd10cm_codes_2026")
PFS_FILE     = os.path.join(_DATA_DIR, "PFREV26B.txt")
PPRRVU_FILE  = os.path.join(_DATA_DIR, "PPRRVU2026_Jan_nonQPP.csv")
GEM_FILE     = os.path.join(_DATA_DIR, "2018_I9gem.txt")


# ── DATABASE SETUP ────────────────────────────────────────────────────────────

def create_tables(conn):
    """Create all reference tables with indexes for fast lookup."""
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS ncci_bundles (
            col1_cpt  TEXT NOT NULL,
            col2_cpt  TEXT NOT NULL,
            modifier  TEXT,
            rationale TEXT,
            PRIMARY KEY (col1_cpt, col2_cpt)
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_ncci_col1 ON ncci_bundles(col1_cpt)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_ncci_col2 ON ncci_bundles(col2_cpt)")

    c.execute("""
        CREATE TABLE IF NOT EXISTS mue_limits (
            cpt_code   TEXT PRIMARY KEY,
            mue_value  INTEGER NOT NULL,
            indicator  TEXT,
            rationale  TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS icd10_codes (
            icd10_code  TEXT PRIMARY KEY,
            description TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS cpt_rates (
            cpt_code    TEXT PRIMARY KEY,
            avg_cost    REAL NOT NULL,
            description TEXT,
            specialty   TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS icd9_crosswalk (
            icd9_code   TEXT PRIMARY KEY,
            icd10_code  TEXT NOT NULL,
            is_exact    INTEGER DEFAULT 0
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_icd9 ON icd9_crosswalk(icd9_code)")

    c.execute("""
        CREATE TABLE IF NOT EXISTS icd_cpt_rules (
            icd10_code   TEXT NOT NULL,
            cpt_code     TEXT NOT NULL,
            rule_source  TEXT DEFAULT 'domain_curated',
            notes        TEXT,
            PRIMARY KEY (icd10_code, cpt_code)
        )
    """)
    c.execute("CREATE INDEX IF NOT EXISTS idx_icd_rules ON icd_cpt_rules(icd10_code)")

    # Metadata table — tracks when each table was last loaded
    c.execute("""
        CREATE TABLE IF NOT EXISTS db_metadata (
            table_name   TEXT PRIMARY KEY,
            loaded_at    TEXT,
            record_count INTEGER,
            source_file  TEXT
        )
    """)

    conn.commit()
    print("  ✅ Tables created")


# ── LOAD NCCI ─────────────────────────────────────────────────────────────────

def load_ncci(conn):
    if not os.path.exists(NCCI_FILE):
        print(f"  ⚠️  NCCI file not found — skipping")
        return 0

    print(f"  [NCCI] Loading from {os.path.basename(NCCI_FILE)}...")
    t = time.time()

    wb = openpyxl.load_workbook(NCCI_FILE, read_only=True, data_only=True)
    ws = wb.active

    rows    = []
    skipped = 0

    for row in ws.iter_rows(min_row=7, values_only=True):
        col1     = str(row[0]).strip() if row[0] else ""
        col2     = str(row[1]).strip() if row[1] else ""
        deletion = str(row[4]).strip() if row[4] else ""
        modifier = str(row[5]).strip() if row[5] else ""
        rationale= str(row[6]).strip() if row[6] else ""

        if not col1 or not col2:
            skipped += 1
            continue
        if deletion != "*":
            skipped += 1
            continue
        if modifier != "0":
            skipped += 1
            continue

        rows.append((
            col1.zfill(5),
            col2.zfill(5),
            modifier,
            rationale
        ))

    wb.close()

    c = conn.cursor()
    c.execute("DELETE FROM ncci_bundles")
    c.executemany(
        "INSERT OR REPLACE INTO ncci_bundles (col1_cpt, col2_cpt, modifier, rationale) VALUES (?,?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("ncci_bundles", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(NCCI_FILE))
    )
    conn.commit()

    print(f"  [NCCI] ✅ {len(rows):,} pairs loaded ({time.time()-t:.1f}s)")
    return len(rows)


# ── LOAD MUE ──────────────────────────────────────────────────────────────────

def load_mue(conn):
    if not os.path.exists(MUE_FILE):
        print(f"  ⚠️  MUE file not found — skipping")
        return 0

    print(f"  [MUE]  Loading from {os.path.basename(MUE_FILE)}...")
    t = time.time()

    wb = openpyxl.load_workbook(MUE_FILE, read_only=True, data_only=True)
    ws = wb.active

    rows    = []
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        cpt_code  = str(row[0]).strip() if row[0] else ""
        mue_value = row[1]
        indicator = str(row[2]).strip() if row[2] else ""
        rationale = str(row[3]).strip() if row[3] else ""

        if not cpt_code or mue_value is None:
            skipped += 1
            continue
        try:
            mue_int = int(float(str(mue_value)))
            if mue_int <= 0:
                skipped += 1
                continue
        except (ValueError, TypeError):
            skipped += 1
            continue

        rows.append((cpt_code, mue_int, indicator, rationale))

    wb.close()

    c = conn.cursor()
    c.execute("DELETE FROM mue_limits")
    c.executemany(
        "INSERT OR REPLACE INTO mue_limits (cpt_code, mue_value, indicator, rationale) VALUES (?,?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("mue_limits", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(MUE_FILE))
    )
    conn.commit()

    print(f"  [MUE]  ✅ {len(rows):,} CPT limits loaded ({time.time()-t:.1f}s)")
    return len(rows)


# ── LOAD ICD-10-CM ────────────────────────────────────────────────────────────

def load_icd10(conn):
    filepath = ICD10_FILE
    if not os.path.exists(filepath):
        filepath = ICD10_FILE + ".txt"
    if not os.path.exists(filepath):
        print(f"  ⚠️  ICD-10-CM file not found — skipping")
        return 0

    print(f"  [ICD10] Loading from {os.path.basename(filepath)}...")
    t = time.time()

    rows    = []
    skipped = 0

    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                skipped += 1
                continue
            parts = line.split(None, 1)
            if len(parts) < 2:
                skipped += 1
                continue
            raw_code    = parts[0].strip()
            description = parts[1].strip()
            if not raw_code or not description:
                skipped += 1
                continue
            # Add dot after 3rd character
            if len(raw_code) > 3:
                formatted = raw_code[:3] + "." + raw_code[3:]
            else:
                formatted = raw_code
            rows.append((formatted, description))

    c = conn.cursor()
    c.execute("DELETE FROM icd10_codes")
    c.executemany(
        "INSERT OR REPLACE INTO icd10_codes (icd10_code, description) VALUES (?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("icd10_codes", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(filepath))
    )
    conn.commit()

    print(f"  [ICD10] ✅ {len(rows):,} ICD-10 codes loaded ({time.time()-t:.1f}s)")
    return len(rows)


# ── LOAD PHYSICIAN FEE SCHEDULE ───────────────────────────────────────────────

def load_pfs(conn):
    if not os.path.exists(PFS_FILE):
        print(f"  ⚠️  PFS file not found — skipping")
        return 0

    print(f"  [PFS]  Loading from {os.path.basename(PFS_FILE)}...")
    t = time.time()

    # Known CPT descriptions and specialties for our key codes
    known_descriptions = {
        "99213": ("Office visit low-moderate complexity",  ["Family Practice", "Internal Medicine"]),
        "99214": ("Office visit moderate complexity",       ["Internal Medicine", "Family Practice"]),
        "99215": ("Office visit high complexity",           ["Internal Medicine"]),
        "36415": ("Venipuncture",                           ["Family Practice", "Internal Medicine"]),
        "80053": ("Comprehensive metabolic panel",          ["Internal Medicine"]),
        "93000": ("Electrocardiogram with interpretation",  ["Cardiology"]),
        "27447": ("Total knee replacement arthroplasty",    ["Orthopedic Surgery"]),
        "45378": ("Colonoscopy diagnostic",                 ["Gastroenterology"]),
        "71046": ("Chest X-ray 2 views",                   ["Internal Medicine", "Radiology"]),
        "99232": ("Inpatient hospital care subsequent",     ["Internal Medicine"]),
        "00100": ("Anesthesia for salivary gland surgery",  ["Anesthesiology"]),
    }

    # Accumulate all locality rates per CPT, then average — this file has no
    # national (00000) rows; it is locality-only, so averaging gives the
    # national representative PFS benchmark rate.
    rate_buckets: dict[str, list[float]] = {}

    with open(PFS_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = [p.strip().strip('"') for p in line.split(",")]
            if len(parts) < 6:
                continue

            cpt_code = parts[3].strip().zfill(5) if parts[3].strip() else ""
            modifier = parts[4].strip()
            amount   = parts[5].strip()

            if not cpt_code:
                continue
            # Skip modifier rows — only unmodified (blank) rates are base PFS rates
            if modifier and modifier.strip():
                continue

            try:
                rate = float(amount.lstrip("0") or "0")
            except (ValueError, TypeError):
                continue

            if rate > 0:
                rate_buckets.setdefault(cpt_code, []).append(rate)

    rows = []
    for cpt_code, rates in rate_buckets.items():
        avg_cost = round(sum(rates) / len(rates), 2)
        desc, specialty = known_descriptions.get(cpt_code, ("", []))
        rows.append((cpt_code, avg_cost, desc, ",".join(specialty)))

    c = conn.cursor()
    c.execute("DELETE FROM cpt_rates")
    c.executemany(
        "INSERT OR REPLACE INTO cpt_rates (cpt_code, avg_cost, description, specialty) VALUES (?,?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("cpt_rates", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(PFS_FILE))
    )
    conn.commit()

    print(f"  [PFS]  ✅ {len(rows):,} CPT rates loaded ({time.time()-t:.1f}s)")
    return len(rows)


# ── LOAD PFS FROM PPRRVU (full 10,000+ code RVU file) ────────────────────────

def load_pfs_from_pprrvu(conn):
    """
    Load CMS Physician Fee Schedule rates from PPRRVU2026_Jan_nonQPP.csv.

    Column positions (0-based, after skipping the 9-row header block):
      0  = HCPCS code
      2  = Description
      3  = Status code  (only 'A' = active, separately payable)
      5  = Work RVU
      6  = Non-Facility PE RVU
      10 = Malpractice (MP) RVU
      25 = Conversion Factor

    Payment formula:
      payment = (Work_RVU + NonFac_PE_RVU + MP_RVU) × Conversion_Factor
    """
    if not os.path.exists(PPRRVU_FILE):
        print(f"  ⚠️  PPRRVU file not found — skipping ({PPRRVU_FILE})")
        return 0

    print(f"  [PFS2] Loading from {os.path.basename(PPRRVU_FILE)}...")
    t = time.time()

    rows    = []
    skipped = 0

    with open(PPRRVU_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line_num, line in enumerate(f):
            # Skip the 10-row header block (rows 0-9)
            if line_num < 10:
                continue

            parts = [p.strip().strip('"') for p in line.split(",")]

            # Need at least 26 columns to reach the conversion factor
            if len(parts) < 26:
                skipped += 1
                continue

            cpt_code    = parts[0].strip().zfill(5) if parts[0].strip() else ""
            description = parts[2].strip()
            status_code = parts[3].strip()

            # Only load Status A = active, separately payable procedures
            if status_code != "A":
                skipped += 1
                continue

            if not cpt_code:
                skipped += 1
                continue

            # Parse RVU fields — skip rows with "NA", blank, or non-numeric values
            try:
                work_rvu  = parts[5].strip()
                nonfac_pe = parts[6].strip()
                mp_rvu    = parts[10].strip()
                cf        = parts[25].strip()

                if any(v in ("", "NA", "N/A") for v in (work_rvu, nonfac_pe, mp_rvu, cf)):
                    skipped += 1
                    continue

                work_rvu  = float(work_rvu)
                nonfac_pe = float(nonfac_pe)
                mp_rvu    = float(mp_rvu)
                cf        = float(cf)
            except (ValueError, TypeError):
                skipped += 1
                continue

            if cf == 0:
                skipped += 1
                continue

            payment = round((work_rvu + nonfac_pe + mp_rvu) * cf, 2)

            rows.append((cpt_code, payment, description, ""))

    c = conn.cursor()
    c.execute("DELETE FROM cpt_rates")
    c.executemany(
        "INSERT OR REPLACE INTO cpt_rates (cpt_code, avg_cost, description, specialty) VALUES (?,?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("cpt_rates", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(PPRRVU_FILE))
    )
    conn.commit()

    print(f"  [PFS2] ✅ {len(rows):,} CPT rates loaded, {skipped:,} skipped ({time.time()-t:.1f}s)")

    # Spot-check key CPT codes
    spot_codes = ("99215", "27447", "36415", "00100")
    print(f"  [PFS2] Spot check:")
    for code in spot_codes:
        row = conn.execute(
            "SELECT cpt_code, avg_cost, description FROM cpt_rates WHERE cpt_code = ?",
            (code.zfill(5),)
        ).fetchone()
        if row:
            print(f"           CPT {row[0]}  ${row[1]:>10,.2f}  {row[2]}")
        else:
            print(f"           CPT {code}  NOT FOUND")

    return len(rows)


# ── LOAD ICD-9 CROSSWALK ──────────────────────────────────────────────────────

def load_gem(conn):
    if not os.path.exists(GEM_FILE):
        print(f"  ⚠️  GEM crosswalk not found — skipping")
        return 0

    print(f"  [GEM]  Loading from {os.path.basename(GEM_FILE)}...")
    t = time.time()

    icd9_map = {}

    with open(GEM_FILE, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) < 3:
                continue
            icd9  = parts[0].strip()
            icd10 = parts[1].strip()
            flags = parts[2].strip()

            # Skip no-map entries
            if len(flags) >= 2 and flags[1] == "1":
                continue

            is_exact = 1 if (len(flags) >= 1 and flags[0] == "0") else 0

            if icd9 not in icd9_map:
                icd9_map[icd9] = (icd10, is_exact)
            else:
                # Prefer exact maps
                if is_exact > icd9_map[icd9][1]:
                    icd9_map[icd9] = (icd10, is_exact)

    rows = [(icd9, icd10, is_exact) for icd9, (icd10, is_exact) in icd9_map.items()]

    c = conn.cursor()
    c.execute("DELETE FROM icd9_crosswalk")
    c.executemany(
        "INSERT OR REPLACE INTO icd9_crosswalk (icd9_code, icd10_code, is_exact) VALUES (?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("icd9_crosswalk", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), os.path.basename(GEM_FILE))
    )
    conn.commit()

    print(f"  [GEM]  ✅ {len(rows):,} ICD-9→ICD-10 mappings loaded ({time.time()-t:.1f}s)")
    return len(rows)


# ── LOAD ICD-CPT VALIDATION RULES ────────────────────────────────────────────

def load_icd_cpt_rules(conn):
    """
    Load domain-curated ICD-10 → valid CPT mappings into icd_cpt_rules table.

    WHY THIS EXISTS:
    CMS does not publish an ICD-to-CPT mapping file. These mappings are
    clinical domain knowledge derived from:
      - AMA CPT coding guidelines
      - CMS Medicare coverage policies (NCDs/LCDs)
      - Medical necessity criteria by specialty
      - Clinical practice standards

    FUTURE EXPANSION:
      - Mine SynPUF co-occurrence patterns (Option C from architecture discussion)
      - Parse CMS LCD policy files (Option B)
      - Add specialty-specific mappings as pipeline matures

    Each row = one valid ICD-CPT pair. If an ICD code has N valid CPTs,
    there are N rows for that ICD code.
    """

    # Domain-curated ICD-10 → valid CPT mappings
    # Source: Clinical coding guidelines + medical necessity criteria
    ICD_CPT_MAPPINGS = [
        # Orthopedic / musculoskeletal
        ("M17.11", "27447", "Knee OA → total knee replacement"),
        ("M17.11", "29881", "Knee OA → knee arthroscopy"),
        ("M17.11", "97110", "Knee OA → therapeutic exercise"),
        ("M17.12", "27447", "Knee OA left → total knee replacement"),
        ("M17.12", "29881", "Knee OA left → knee arthroscopy"),
        ("M17.12", "97110", "Knee OA left → therapeutic exercise"),
        ("M16.11", "27130", "Hip OA right → total hip replacement"),
        # ICD_CPT_MISMATCH fraud scenario: hip stiffness billed with knee surgery
        ("M25.361","97110", "Hip stiffness → therapeutic exercise"),
        ("M25.361","97035", "Hip stiffness → ultrasound therapy"),
        ("M25.361","99213", "Hip stiffness → office visit"),
        # NOTE: 27447 (total knee replacement) is intentionally NOT listed
        # for M25.361 (hip stiffness) — this is our injected ICD_CPT_MISMATCH scenario

        # Cardiovascular
        ("I10",    "99213", "Hypertension → office visit low"),
        ("I10",    "99214", "Hypertension → office visit moderate"),
        ("I10",    "93000", "Hypertension → EKG"),
        ("I10",    "80053", "Hypertension → metabolic panel"),
        ("I25.10", "93000", "CAD → EKG"),
        ("I25.10", "93306", "CAD → echocardiogram"),
        ("I25.10", "99214", "CAD → office visit moderate"),
        ("I25.10", "99215", "CAD → office visit high"),
        ("R00.0",  "93000", "Tachycardia → EKG"),
        ("R00.0",  "99213", "Tachycardia → office visit low"),
        ("R00.0",  "99214", "Tachycardia → office visit moderate"),

        # Gastrointestinal
        ("Z12.11", "45378", "Colorectal cancer screening → diagnostic colonoscopy"),
        ("Z12.11", "45380", "Colorectal cancer screening → colonoscopy with biopsy"),
        ("K57.30", "45378", "Diverticulosis → diagnostic colonoscopy"),
        ("K57.30", "45380", "Diverticulosis → colonoscopy with biopsy"),
        ("K11.5",  "00100", "Sialolithiasis → anesthesia salivary gland"),
        ("K11.5",  "42330", "Sialolithiasis → submandibular gland stone removal"),

        # Neurology
        ("G43.909","70553", "Migraine → MRI brain with contrast"),
        ("G43.909","99214", "Migraine → office visit moderate"),
        ("G43.909","99215", "Migraine → office visit high"),

        # Psychiatry
        ("F32.1",  "90837", "MDD moderate → psychotherapy 60 min"),
        ("F32.1",  "99214", "MDD moderate → office visit moderate"),

        # Endocrinology / metabolic
        ("E11.9",  "99213", "T2DM → office visit low"),
        ("E11.9",  "99214", "T2DM → office visit moderate"),
        ("E11.9",  "80053", "T2DM → comprehensive metabolic panel"),
        ("E11.9",  "85025", "T2DM → CBC"),
        ("E11.9",  "36415", "T2DM → venipuncture"),

        # Respiratory / infectious
        ("J06.9",  "99212", "URTI → office visit minimal"),
        ("J06.9",  "99213", "URTI → office visit low"),
        # NOTE: 99215 (high complexity) is intentionally NOT listed for J06.9
        # This is our UPCODING fraud scenario

        ("J18.9",  "71046", "Pneumonia → chest X-ray 2 views"),
        ("J18.9",  "99213", "Pneumonia → office visit low"),
        ("J18.9",  "99214", "Pneumonia → office visit moderate"),
        ("J18.9",  "99232", "Pneumonia → subsequent inpatient care"),
        ("J18.9",  "99233", "Pneumonia → subsequent inpatient care high"),

        # Genitourinary
        ("N39.0",  "99213", "UTI → office visit low"),
        ("N39.0",  "99214", "UTI → office visit moderate"),
        ("N39.0",  "99232", "UTI → subsequent inpatient care"),
        ("N39.0",  "87088", "UTI → urine culture"),
        ("N39.0",  "81001", "UTI → urinalysis"),

        # Oncology / pulmonary
        ("C34.11", "99215", "Lung cancer → office visit high"),
        ("C34.11", "99214", "Lung cancer → office visit moderate"),
        ("C34.11", "71046", "Lung cancer → chest X-ray 2 views"),
        ("C34.11", "71250", "Lung cancer → chest CT without contrast"),

        # Preventive / wellness
        # NOTE: 36415 (venipuncture) is intentionally NOT listed for Z00.00
        # This is our MEDICALLY_UNNECESSARY fraud scenario
        ("Z00.00", "99395", "Wellness exam → preventive medicine 18-39"),
        ("Z00.00", "99396", "Wellness exam → preventive medicine 40-64"),
        ("Z00.00", "99397", "Wellness exam → preventive medicine 65+"),
    ]

    rows = [(icd, cpt, "domain_curated", note)
            for icd, cpt, note in ICD_CPT_MAPPINGS]

    t = time.time()
    print(f"  [ICD→CPT] Loading domain-curated ICD-CPT validation rules...")

    c = conn.cursor()
    c.execute("DELETE FROM icd_cpt_rules")
    c.executemany(
        "INSERT OR REPLACE INTO icd_cpt_rules (icd10_code, cpt_code, rule_source, notes) VALUES (?,?,?,?)",
        rows
    )
    c.execute(
        "INSERT OR REPLACE INTO db_metadata VALUES (?,?,?,?)",
        ("icd_cpt_rules", time.strftime("%Y-%m-%d %H:%M:%S"), len(rows), "domain_curated")
    )
    conn.commit()

    print(f"  [ICD→CPT] ✅ {len(rows):,} ICD→CPT rules loaded ({time.time()-t:.1f}s)")
    print(f"  [ICD→CPT]    Covers {len(set(r[0] for r in rows)):,} unique ICD codes")
    print(f"  [ICD→CPT]    FWA scenarios preserved: UPCODING, UNBUNDLING, ICD_CPT_MISMATCH, MEDICALLY_UNNECESSARY")
    return len(rows)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def build_database():
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  MediGuard AI — CMS Reference Database Setup                     ║")
    print("╚══════════════════════════════════════════════════════════════════╝")
    print()
    print(f"  Database: {DB_PATH}")
    print()

    os.makedirs(_DATA_DIR, exist_ok=True)
    t_total = time.time()

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    print("  Creating tables...")
    create_tables(conn)
    print()

    print("  Loading CMS reference data...")
    print("  ─────────────────────────────────────────────────────────────")
    n_ncci  = load_ncci(conn)
    n_mue   = load_mue(conn)
    n_icd10 = load_icd10(conn)
    n_pfs   = load_pfs_from_pprrvu(conn)
    n_gem   = load_gem(conn)
    n_rules = load_icd_cpt_rules(conn)

    conn.close()

    print()
    print("  ═══════════════════════════════════════════════════════════════")
    print("  DATABASE BUILD COMPLETE")
    print("  ═══════════════════════════════════════════════════════════════")
    print(f"  Total time          : {time.time()-t_total:.1f}s")
    print(f"  Database location   : {DB_PATH}")
    db_size = os.path.getsize(DB_PATH) / (1024*1024) if os.path.exists(DB_PATH) else 0
    print(f"  Database size       : {db_size:.1f} MB")
    print()
    print(f"  Tables loaded:")
    print(f"    ncci_bundles      : {n_ncci:>8,} rows")
    print(f"    mue_limits        : {n_mue:>8,} rows")
    print(f"    icd10_codes       : {n_icd10:>8,} rows")
    print(f"    cpt_rates         : {n_pfs:>8,} rows")
    print(f"    icd9_crosswalk    : {n_gem:>8,} rows")
    print(f"    icd_cpt_rules     : {n_rules:>8,} rows")
    print()
    print("  Next step: Run fwa_data_pipeline.py — it will use the DB automatically")
    print()


if __name__ == "__main__":
    build_database()
