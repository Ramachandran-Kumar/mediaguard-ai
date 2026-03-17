"""
╔══════════════════════════════════════════════════════════════════╗
║  MEDIAGUARD AI — CMS Reference Data Loader                       ║
║  Replaces hardcoded NCCI_BUNDLES and MUE_LIMITS with real CMS data║
╚══════════════════════════════════════════════════════════════════╝

PURPOSE:
  Loads two official CMS reference files and returns them in the
  exact format that fwa_data_pipeline.py expects — so the pipeline
  needs zero changes. Just swap the hardcoded dicts for these loaders.

FILES REQUIRED (save in mediaguard-ai/data/):
  1. ccipra-v321r0-f1.xlsx         ← NCCI PTP Practitioner edits
  2. MCR_MUE_PractitionerServices_Eff_04-01-2026.xlsx  ← MUE limits

WHAT REPLACES WHAT:
  load_ncci_bundles()  →  replaces hardcoded NCCI_BUNDLES  (was 6 pairs)
  load_mue_limits()    →  replaces hardcoded MUE_LIMITS    (was 7 entries)

USAGE IN fwa_data_pipeline.py:
  # Replace the hardcoded dicts at the top of the file with:
  from cms_reference_loader import load_ncci_bundles, load_mue_limits
  NCCI_BUNDLES = load_ncci_bundles()
  MUE_LIMITS   = load_mue_limits()

NCCI FILE STRUCTURE (ccipra-v321r0-f1.xlsx):
  Headers at row 3: Column 1, Column 2, *=in existence, Effective,
                    Deletion, Modifier, PTP Edit Rationale
  Data starts row 7
  Active edits identified by: Deletion == '*'
  Modifier: 0 = hard edit (modifier not allowed)
             1 = soft edit (modifier allowed)
             9 = not applicable

MUE FILE STRUCTURE (MCR_MUE_PractitionerServices_Eff_04-01-2026.xlsx):
  Headers at row 2: HCPCS/CPT Code, Practitioner Services MUE Values,
                    MUE Adjudication Indicator, MUE Rationale
  Data starts row 3
"""

import os
import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────────────
# Paths relative to project root — adjust if your data folder is elsewhere
_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

NCCI_FILE = os.path.join(_DATA_DIR, "ccipra-v321r0-f1.xlsx")
MUE_FILE  = os.path.join(_DATA_DIR, "MCR_MUE_PractitionerServices_Eff_04-01-2026.xlsx")

# ── FALLBACK HARDCODED VALUES ─────────────────────────────────────────────────
# Used if the CMS files are not found — keeps pipeline runnable without the files
# These are the original hardcoded values from fwa_data_pipeline.py

_FALLBACK_NCCI_BUNDLES = {
    ("36415", "36416"),   # Venipuncture + capillary blood draw
    ("99213", "99201"),   # Office visits — lower cannot be billed with higher
    ("99214", "99201"),
    ("99214", "99202"),
    ("99215", "99201"),
    ("99215", "99202"),
}

_FALLBACK_MUE_LIMITS = {
    "99213": 1,
    "99214": 1,
    "99215": 1,
    "36415": 1,
    "80053": 1,
    "93000": 1,
    "27447": 1,
}


# ── NCCI LOADER ───────────────────────────────────────────────────────────────

def load_ncci_bundles(modifier_filter="hard_only"):
    """
    Load CMS NCCI PTP edits and return as a set of (col1_cpt, col2_cpt) tuples.

    These are the pairs that CANNOT be billed together on the same day
    for the same patient by the same provider.

    Args:
        modifier_filter:
            "hard_only"  — Only include modifier=0 edits (modifier NOT allowed).
                           These are the strictest rules — a modifier cannot
                           override them. Best for FWA detection (fewer FP).
            "all_active" — Include modifier=0 AND modifier=1 edits.
                           Modifier=1 means a modifier CAN override the rule,
                           but billers sometimes skip the modifier intentionally.
            "soft_only"  — Only modifier=1 edits (soft edits).

    Returns:
        set of tuples: {("36415", "36416"), ("99213", "99212"), ...}
        Falls back to hardcoded values if file not found.

    Pipeline usage:
        NCCI_BUNDLES = load_ncci_bundles()
        # Then in rule engine:
        # if (cpt1, cpt2) in NCCI_BUNDLES: flag as unbundling
    """
    if not os.path.exists(NCCI_FILE):
        print(f"  ⚠️  NCCI file not found at: {NCCI_FILE}")
        print(f"      Using {len(_FALLBACK_NCCI_BUNDLES)} hardcoded fallback pairs.")
        print(f"      Download from: cms.gov/medicare/coding-billing/national-correct-coding-initiative-edits")
        return _FALLBACK_NCCI_BUNDLES

    print(f"  [NCCI] Loading CMS NCCI PTP edits from: {os.path.basename(NCCI_FILE)}")

    try:
        wb = openpyxl.load_workbook(NCCI_FILE, read_only=True, data_only=True)
        ws = wb.active

        bundles = set()
        skipped  = 0
        loaded   = 0

        for row in ws.iter_rows(min_row=7, values_only=True):
            col1     = row[0]
            col2     = row[1]
            deletion = row[4]   # '*' = still active; a date = expired
            modifier = str(row[5]).strip() if row[5] is not None else ""

            # Skip header/legend rows and blank rows
            if col1 is None or col2 is None:
                skipped += 1
                continue
            if not str(col1).strip() or not str(col2).strip():
                skipped += 1
                continue

            # Only load ACTIVE edits — deletion date must be '*'
            if str(deletion).strip() != "*":
                skipped += 1
                continue

            # Apply modifier filter
            if modifier_filter == "hard_only" and modifier != "0":
                skipped += 1
                continue
            elif modifier_filter == "soft_only" and modifier != "1":
                skipped += 1
                continue
            # "all_active" passes both 0 and 1 (skips 9=not applicable)
            elif modifier_filter == "all_active" and modifier == "9":
                skipped += 1
                continue

            cpt1 = str(col1).strip().zfill(5)
            cpt2 = str(col2).strip().zfill(5)
            bundles.add((cpt1, cpt2))
            loaded += 1

        wb.close()

        filter_label = {
            "hard_only":  "hard edits only (modifier=0)",
            "all_active": "all active edits (modifier 0+1)",
            "soft_only":  "soft edits only (modifier=1)",
        }.get(modifier_filter, modifier_filter)

        print(f"  [NCCI] ✅ Loaded {loaded:,} active bundling pairs ({filter_label})")
        print(f"  [NCCI]    Skipped {skipped:,} expired or filtered rows")
        print(f"  [NCCI]    Replaces 6 hardcoded pairs → {loaded:,} real CMS pairs")

        return bundles

    except Exception as e:
        print(f"  ❌ Error loading NCCI file: {e}")
        print(f"     Falling back to {len(_FALLBACK_NCCI_BUNDLES)} hardcoded pairs.")
        return _FALLBACK_NCCI_BUNDLES


# ── MUE LOADER ────────────────────────────────────────────────────────────────

def load_mue_limits():
    """
    Load CMS Medically Unlikely Edits (MUE) and return as a dict.

    MUE limits define the maximum units of a service that can be billed
    per day per patient by one provider. Billing above this limit is
    medically unlikely and triggers a flag.

    Returns:
        dict: {"99215": 1, "36415": 1, "27447": 2, ...}
              CPT code (str) → max units per day (int)
        Falls back to hardcoded values if file not found.

    Pipeline usage:
        MUE_LIMITS = load_mue_limits()
        # Then in rule engine:
        # if units_billed > MUE_LIMITS.get(cpt_code, 999): flag as MUE violation
    """
    if not os.path.exists(MUE_FILE):
        print(f"  ⚠️  MUE file not found at: {MUE_FILE}")
        print(f"      Using {len(_FALLBACK_MUE_LIMITS)} hardcoded fallback entries.")
        print(f"      Download from: cms.gov/medicare/coding-billing/medicare-edits/medically-unlikely-edits")
        return _FALLBACK_MUE_LIMITS

    print(f"  [MUE]  Loading CMS MUE limits from: {os.path.basename(MUE_FILE)}")

    try:
        wb = openpyxl.load_workbook(MUE_FILE, read_only=True, data_only=True)
        ws = wb.active

        mue_limits = {}
        skipped    = 0
        loaded     = 0

        # Headers are at row 2, data starts at row 3
        for row in ws.iter_rows(min_row=3, values_only=True):
            cpt_code  = row[0]
            mue_value = row[1]

            # Skip blank rows
            if cpt_code is None or mue_value is None:
                skipped += 1
                continue

            cpt_str = str(cpt_code).strip()
            if not cpt_str:
                skipped += 1
                continue

            # MUE value should be a positive integer
            try:
                mue_int = int(float(str(mue_value).strip()))
                if mue_int <= 0:
                    skipped += 1
                    continue
            except (ValueError, TypeError):
                skipped += 1
                continue

            mue_limits[cpt_str] = mue_int
            loaded += 1

        wb.close()

        print(f"  [MUE]  ✅ Loaded {loaded:,} CPT MUE limits")
        print(f"  [MUE]     Skipped {skipped:,} blank or invalid rows")
        print(f"  [MUE]     Replaces 7 hardcoded entries → {loaded:,} real CMS entries")

        return mue_limits

    except Exception as e:
        print(f"  ❌ Error loading MUE file: {e}")
        print(f"     Falling back to {len(_FALLBACK_MUE_LIMITS)} hardcoded entries.")
        return _FALLBACK_MUE_LIMITS


# ── SUMMARY HELPER ────────────────────────────────────────────────────────────

def load_all_reference_data(ncci_modifier_filter="hard_only"):
    """
    Load both CMS reference tables in one call.
    Prints a clean summary for pipeline startup logs.

    Returns:
        tuple: (ncci_bundles set, mue_limits dict)

    Usage:
        NCCI_BUNDLES, MUE_LIMITS = load_all_reference_data()
    """
    print()
    print("  ── CMS Reference Data Loader ─────────────────────────────────")
    ncci = load_ncci_bundles(modifier_filter=ncci_modifier_filter)
    mue  = load_mue_limits()
    print("  ── Reference data ready ──────────────────────────────────────")
    print()
    return ncci, mue


# ── QUICK TEST ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║  MediGuard AI — CMS Reference Loader Test                        ║")
    print("╚══════════════════════════════════════════════════════════════════╝")

    ncci_bundles, mue_limits = load_all_reference_data()

    # Spot check — our known NCCI pair
    test_pair = ("36415", "36416")
    print(f"\n  Spot check NCCI — is {test_pair} in loaded bundles?")
    print(f"    {'✅ YES' if test_pair in ncci_bundles else '❌ NOT FOUND'}")

    # Spot check — our known MUE entry
    test_cpt = "99215"
    print(f"\n  Spot check MUE — what is the limit for CPT {test_cpt}?")
    print(f"    Limit: {mue_limits.get(test_cpt, 'NOT FOUND')}")

    # Sample 5 NCCI pairs
    print(f"\n  Sample NCCI pairs (first 5):")
    for pair in list(ncci_bundles)[:5]:
        print(f"    {pair[0]} + {pair[1]}")

    # Sample 5 MUE entries
    print(f"\n  Sample MUE limits (first 5):")
    for cpt, limit in list(mue_limits.items())[:5]:
        print(f"    CPT {cpt} → max {limit} unit(s)/day")

    print("\n  ✅ CMS reference loader test complete.")
    print("  Ready to import into fwa_data_pipeline.py")
